import requests
import io
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from fpdf import FPDF
from textwrap import wrap
import os
import time

# 🔍 Chargement des données SCImago depuis le CSV - VERSION AMÉLIORÉE
def load_scimago_data(csv_path):
    scimago_info = {}
    with open(csv_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            title = row["Title"].strip().lower()
            
            journal_data = {
                "issn": row["Issn"],
                "sjr": row["SJR"],
                "quartile": row["SJR Best Quartile"],
                "h_index": row["H index"]
            }
            
            # Stocker le titre exact
            scimago_info[title] = journal_data
            
            # Créer des variantes pour améliorer le matching
            # Variante sans ponctuation
            title_no_punct = re.sub(r'[:\-,;\.&\(\)\[\]_]+', ' ', title)
            title_no_punct = ' '.join(title_no_punct.split())
            if title_no_punct != title:
                scimago_info[title_no_punct] = journal_data
            
            # Variantes supplémentaires
            variants = [
                title.replace(':', ' '),
                title.replace('-', ' '),
                title.replace('&', 'and'),
                title.replace(' and ', ' & '),
                title.replace(',', ''),
                title.replace(';', ''),
                title.replace('_', ' '),
                title.replace('(', '').replace(')', ''),
                title.replace('[', '').replace(']', ''),
            ]
            
            for variant in variants:
                variant = ' '.join(variant.split())
                if variant != title and variant not in scimago_info:
                    scimago_info[variant] = journal_data
            
            # Variantes pour proceedings et conférences
            proceedings_variants = []
            
            # Suppression des mots proceedings
            if 'proceedings' in title:
                clean_title = re.sub(r'\bproceedings?\s+of\s+the\s+', '', title)
                clean_title = re.sub(r'\bproceedings?\s+', '', clean_title)
                clean_title = ' '.join(clean_title.split())
                proceedings_variants.append(clean_title)
            
            # Suppression des numéros d'édition et années
            clean_title = re.sub(r'\b\d+(?:st|nd|rd|th)\s+', '', title)
            clean_title = re.sub(r'\b\d{4}\b', '', clean_title)
            clean_title = re.sub(r'\binternational\s+conference\s+on\s+', '', clean_title)
            clean_title = re.sub(r'\bconference\s+on\s+', '', clean_title)
            clean_title = ' '.join(clean_title.split())
            if clean_title != title:
                proceedings_variants.append(clean_title)
            
            # Ajouter les variantes proceedings
            for variant in proceedings_variants:
                if variant and variant not in scimago_info:
                    scimago_info[variant] = journal_data
                    
    return scimago_info

# 🔍 NOUVELLE FONCTION - Matching intelligent des journaux
def improved_journal_matching(journal_name, scimago_data):
    """
    Matching intelligent des journaux avec gestion spéciale des proceedings
    """
    if not journal_name:
        return {}
    
    original_name = journal_name.strip()
    journal_key = original_name.lower()
    
    # 1. Recherche exacte
    if journal_key in scimago_data:
        return scimago_data[journal_key]
    
    # 2. Recherche avec variantes basiques
    basic_variants = [
        journal_key.replace(':', ' '),
        journal_key.replace('-', ' '),
        journal_key.replace('&', 'and'),
        journal_key.replace(' and ', ' & '),
        journal_key.replace(',', ''),
        journal_key.replace(';', ''),
        journal_key.replace('_', ' '),
        journal_key.replace('(', '').replace(')', ''),
        journal_key.replace('[', '').replace(']', ''),
    ]
    
    for variant in basic_variants:
        variant = ' '.join(variant.split())
        if variant in scimago_data:
            print(f"✅ Journal trouvé avec variante basique: '{journal_name}' -> '{variant}'")
            return scimago_data[variant]
    
    # 3. Nettoyage spécial pour proceedings et conférences
    cleaned_name = journal_key
    
    # Patterns à supprimer pour les proceedings
    proceedings_patterns = [
        r'\bproceedings?\s+of\s+the\s+',
        r'\bproceedings?\s+',
        r'\b\d+(?:st|nd|rd|th)\s+',  # 10th, 1st, 2nd, etc.
        r'\b\d{4}\b',  # Années (2024, 2023, etc.)
        r'\binternational\s+conference\s+on\s+',
        r'\bconference\s+on\s+',
        r'\bworkshop\s+on\s+',
        r'\bsymposium\s+on\s+',
        r'\bieee\s+',
        r'\bacm\s+',
    ]
    
    for pattern in proceedings_patterns:
        cleaned_name = re.sub(pattern, '', cleaned_name, flags=re.IGNORECASE)
    
    # Nettoyer les espaces multiples et la ponctuation
    cleaned_name = re.sub(r'[:\-,;\.&\(\)\[\]_]+', ' ', cleaned_name)
    cleaned_name = ' '.join(cleaned_name.split())
    
    # 4. Recherche avec nom nettoyé
    if cleaned_name and cleaned_name in scimago_data:
        print(f"✅ Journal trouvé après nettoyage: '{journal_name}' -> '{cleaned_name}'")
        return scimago_data[cleaned_name]
    
    # 5. Recherche par mots-clés significatifs
    if cleaned_name:
        keywords = [word for word in cleaned_name.split() 
                   if len(word) > 2 and word not in ['the', 'and', 'for', 'with', 'of', 'in', 'on', 'at']]
        
        if keywords:
            # Recherche de journaux contenant tous les mots-clés
            for key, data in scimago_data.items():
                if len(keywords) >= 2 and all(keyword in key for keyword in keywords):
                    print(f"✅ Journal trouvé par mots-clés complets: '{journal_name}' -> '{key}'")
                    return data
            
            # Recherche de journaux contenant au moins la moitié des mots-clés
            min_matches = max(1, len(keywords) // 2)
            best_match = None
            best_score = 0
            
            for key, data in scimago_data.items():
                matches = sum(1 for keyword in keywords if keyword in key)
                if matches >= min_matches and matches > best_score:
                    best_match = (key, data)
                    best_score = matches
            
            if best_match:
                print(f"✅ Journal trouvé par correspondance partielle: '{journal_name}' -> '{best_match[0]}' ({best_score}/{len(keywords)} mots)")
                return best_match[1]
    
    # 6. Recherche par acronymes
    acronyms = re.findall(r'\b[A-Z]{2,}\b', original_name)
    for acronym in acronyms:
        acronym_lower = acronym.lower()
        for key, data in scimago_data.items():
            if acronym_lower in key:
                print(f"✅ Journal trouvé par acronyme '{acronym}': '{journal_name}' -> '{key}'")
                return data
    
    # 7. Recherche floue avancée (similarité de chaînes)
    if len(keywords) >= 2:
        for key, data in scimago_data.items():
            # Calculer un score de similarité simple
            key_words = key.split()
            common_words = set(keywords) & set(key_words)
            if len(common_words) >= 2:
                similarity = len(common_words) / max(len(keywords), len(key_words))
                if similarity > 0.4:  # Seuil de similarité
                    print(f"✅ Journal trouvé par similarité: '{journal_name}' -> '{key}' (score: {similarity:.2f})")
                    return data
    
    print(f"❌ Journal non trouvé: '{journal_name}'")
    return {}

#  NOUVELLE FONCTION - Débogage des journaux
def debug_journal_matching(journal_name, scimago_data, csv_path):
   
    print(f"\n=== 🔍 DÉBOGAGE POUR: {journal_name} ===")
    
    journal_key = journal_name.strip().lower()
    print(f"Nom normalisé: '{journal_key}'")
    
    # Rechercher des correspondances partielles dans le CSV
    print("\n--- 📋 Recherche dans le CSV ---")
    matches_found = []
    
    try:
        with open(csv_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                csv_title = row["Title"].strip().lower()
                
                # Recherche par mots-clés
                words_in_journal = journal_key.split()
                significant_words = [w for w in words_in_journal if len(w) > 3]
                
                if significant_words:
                    matches = sum(1 for word in significant_words if word in csv_title)
                    if matches >= min(2, len(significant_words)):
                        matches_found.append({
                            'title': row["Title"],
                            'normalized': csv_title,
                            'matches': matches,
                            'total_words': len(significant_words),
                            'issn': row["Issn"],
                            'sjr': row["SJR"]
                        })
        
        if matches_found:
            print("🔍 Correspondances partielles trouvées:")
            matches_found.sort(key=lambda x: x['matches'], reverse=True)
            for i, match in enumerate(matches_found[:5]):  # Top 5
                print(f"  {i+1}. '{match['title']}'")
                print(f"     Correspondances: {match['matches']}/{match['total_words']}")
                print(f"     ISSN: {match['issn']}, SJR: {match['sjr']}")
        else:
            print("❌ Aucune correspondance partielle trouvée")
            
    except Exception as e:
        print(f"Erreur lors de la lecture du CSV: {e}")
    
    print("=" * 50)

# Fonction pour extraire le Scopus ID depuis l'EID ou l'URL
def extract_scopus_id(eid_or_url):
    """Extrait le Scopus ID numérique depuis différents formats"""
    if not eid_or_url:
        return None
    
    if eid_or_url.isdigit():
        return eid_or_url
    
    patterns = [
        r'2-s2\.0-(\d+)',
        r'SCOPUS_ID:(\d+)',
        r'scopus_id/(\d+)',
        r'/(\d{10,})',
        r'(\d{10,})'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, str(eid_or_url))
        if match:
            return match.group(1)
    
    return None

# Fonction pour obtenir tous les résultats (pas de limite à 25)
def get_all_publications(query, headers):
    """Récupère toutes les publications sans limite"""
    all_entries = []
    start = 0
    count = 25

    while True:
        url = f"https://api.elsevier.com/content/search/scopus?query={query}&start={start}&count={count}"
        response = requests.get(url, headers=headers, timeout=10)

        if response.status_code != 200:
            print(f"Erreur API: {response.status_code}")
            if response.status_code == 429:
                print("Rate limit atteint, attente de 5 secondes...")
                time.sleep(5)
                continue
            break

        data = response.json()
        entries = data.get("search-results", {}).get("entry", [])

        if not entries:
            break

        all_entries.extend(entries)

        total_results = int(data.get("search-results", {}).get("opensearch:totalResults", 0))
        start += count

        print(f"Récupéré {len(all_entries)} sur {total_results} publications...")

        if start >= total_results:
            break

        time.sleep(0.2)

    return all_entries

def get_coauthors(pub_data, headers):
    """Extrait TOUS les co-auteurs en traitant tous les formats possibles"""
    authors = set()

    # Méthode 1: Extraction depuis le champ 'author'
    if 'author' in pub_data:
        author_entry = pub_data['author']
        
        if isinstance(author_entry, list):
            for author in author_entry:
                name = extract_author_name(author)
                if name:
                    authors.add(name)
        
        elif isinstance(author_entry, dict):
            name = extract_author_name(author_entry)
            if name:
                authors.add(name)
        
        elif isinstance(author_entry, str):
            extracted = process_author_string(author_entry)
            for name in extracted:
                authors.add(name)

    # Méthode 2: Fallback avec dc:creator
    if not authors and 'dc:creator' in pub_data:
        creator_text = pub_data['dc:creator']
        if creator_text:
            extracted = process_author_string(creator_text)
            for name in extracted:
                authors.add(name)

    # Méthode 3: Abstract Retrieval
    if not authors and 'eid' in pub_data:
        scopus_id = extract_scopus_id(pub_data['eid'])
        if scopus_id:
            try:
                abstract_url = f"https://api.elsevier.com/content/abstract/scopus_id/{scopus_id}"
                response = requests.get(abstract_url, headers=headers, timeout=10)
                
                if response.status_code == 200:
                    abstract_data = response.json()
                    extracted = extract_authors_from_abstract(abstract_data)
                    for name in extracted:
                        authors.add(name)
            except Exception as e:
                pass

    return "; ".join(sorted(authors)) if authors else "N/A"

def extract_author_name(author_dict):
    """Extrait un nom d'auteur depuis un dictionnaire Scopus"""
    if not isinstance(author_dict, dict):
        return None
    
    name_fields = ['indexed-name', 'authname', 'surname']
    for field in name_fields:
        if field in author_dict:
            name = author_dict[field].strip()
            if name:
                if field == 'surname' and ('given-name' in author_dict or 'initials' in author_dict):
                    given = author_dict.get('given-name', author_dict.get('initials', ''))
                    return f"{name}, {given}".strip(', ')
                return name
    
    return None

def process_author_string(author_str):
    """Parse une chaîne d'auteurs"""
    if not author_str:
        return []
    
    normalized = author_str.replace(' and ', '; ').replace(' AND ', '; ')
    normalized = re.sub(r',?\s+et\s+', '; ', normalized, flags=re.IGNORECASE)
    normalized = re.sub(r',\s*', '; ', normalized)
    
    parts = [p.strip() for p in normalized.split(';') if p.strip()]
    
    cleaned = []
    for part in parts:
        clean_part = re.sub(r'\s*\[\d+\]\s*', '', part)
        clean_part = ' '.join(clean_part.split())
        if clean_part:
            cleaned.append(clean_part)
    
    return cleaned

def extract_authors_from_abstract(abstract_data):
    """Extrait les auteurs depuis les données d'abstract"""
    authors = set()
    
    authors_list = abstract_data.get('abstracts-retrieval-response', {}).get('authors', {}).get('author', [])
    if isinstance(authors_list, dict):
        authors_list = [authors_list]
    
    for author in authors_list:
        name = extract_author_name(author)
        if name:
            authors.add(name)
    
    if not authors:
        item = abstract_data.get('abstracts-retrieval-response', {}).get('item', {})
        author_groups = item.get('bibrecord', {}).get('head', {}).get('author-group', [])
        if isinstance(author_groups, dict):
            author_groups = [author_groups]
        
        for group in author_groups:
            group_authors = group.get('author', [])
            if isinstance(group_authors, dict):
                group_authors = [group_authors]
            
            for author in group_authors:
                name = extract_author_name(author)
                if name:
                    authors.add(name)
    
    return authors

def get_citations_count(pub_data, headers):
    """Obtient le nombre de citations avec plusieurs méthodes de fallback"""
    
    citations_from_search = pub_data.get("citedby-count", "")
    if citations_from_search and citations_from_search != "0":
        return citations_from_search
    
    scopus_eid = pub_data.get("eid", "")
    scopus_id = extract_scopus_id(scopus_eid)
    
    if scopus_id:
        try:
            time.sleep(0.1)
            
            abstract_url = f"https://api.elsevier.com/content/abstract/scopus_id/{scopus_id}?view=COMPLETE"
            abstract_response = requests.get(abstract_url, headers=headers, timeout=10)

            if abstract_response.status_code == 429:
                time.sleep(2)
                abstract_response = requests.get(abstract_url, headers=headers, timeout=10)
            
            if abstract_response.status_code == 200:
                abstract_data = abstract_response.json()
                coredata = abstract_data.get("abstracts-retrieval-response", {}).get("coredata", {})
                citations_count = coredata.get("citedby-count", "0")
                return citations_count
                
        except Exception as e:
            pass
    
    return citations_from_search if citations_from_search else "0"

def get_document_type(pub_data):
    """
    Extrait et normalise le type de document depuis les données Scopus
    """
    # Types de documents possibles dans Scopus
    document_types = {
        'ar': 'Article',
        'cp': 'Conference Paper',
        'ch': 'Book Chapter', 
        'bk': 'Book',
        're': 'Review',
        'le': 'Letter',
        'ed': 'Editorial',
        'no': 'Note',
        'sh': 'Short Survey',
        'ip': 'Article in Press',
        'ab': 'Abstract Report',
        'wp': 'Working Paper',
        'cr': 'Conference Review',
        'rp': 'Report',
        'tb': 'Trade Publication',
        'dp': 'Data Paper',
        'er': 'Erratum'
    }
    
    # Méthode 1: Champ subtypeDescription (le plus fiable)
    subtype_desc = pub_data.get("subtypeDescription", "")
    if subtype_desc:
        return subtype_desc
    
    # Méthode 2: Champ subtype avec mapping
    subtype = pub_data.get("subtype", "")
    if subtype and subtype.lower() in document_types:
        return document_types[subtype.lower()]
    
    # Méthode 3: Analyse du nom de publication pour détecter les conférences
    publication_name = pub_data.get("prism:publicationName", "").lower()
    conference_keywords = [
        'proceedings', 'conference', 'symposium', 'workshop', 
        'congress', 'meeting', 'summit', 'forum'
    ]
    
    if any(keyword in publication_name for keyword in conference_keywords):
        return "Conference Paper"
    
    # Méthode 4: Champ aggregationType
    aggregation_type = pub_data.get("prism:aggregationType", "")
    if aggregation_type:
        if aggregation_type.lower() == "journal":
            return "Article"
        elif aggregation_type.lower() == "book":
            return "Book Chapter"
        elif aggregation_type.lower() == "conference proceeding":
            return "Conference Paper"
    
    # Valeur par défaut
    return "Article"

# 📄 Classe PDF personnalisée
class PDF(FPDF):
    def __init__(self):
        super().__init__(orientation='L', unit='mm', format='A4')
        self.set_auto_page_break(auto=True, margin=10)
        try:
            self.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
            self.set_font('DejaVu', size=8)
        except:
            self.set_font('Arial', size=8)

    def add_title(self, author_name, year):
        """Ajoute un titre formaté au début du rapport"""
        try:
            self.set_font('DejaVu', size=16)
        except:
            self.set_font('Arial', size=16)
            
        year_text = year if year else "Toutes les années"
        
        title_line1 = f"Rapport de publications de : {author_name}"
        title_line2 = f"Année : {year_text}"
        
        self.set_xy(0, 20)
        self.cell(0, 10, title_line1, ln=True, align='C')
        self.set_x(0)
        self.cell(0, 10, title_line2, ln=True, align='C')
        self.ln(15)
        
        try:
            self.set_font('DejaVu', size=8)
        except:
            self.set_font('Arial', size=8)

    def render_table(self, data, headers, col_widths, line_height):
        try:
            self.set_font('DejaVu', size=8)
        except:
            self.set_font('Arial', size=8)
        
        for i, header in enumerate(headers):
            self.set_xy(self.get_x(), self.get_y())
            self.cell(col_widths[i], 8, header, border=1, align='C')
            current_x = self.get_x() - col_widths[i]
            current_y = self.get_y()
            self.set_xy(current_x + 0.1, current_y)
            self.cell(col_widths[i], 8, header, border=0, align='C')
            self.set_xy(current_x + col_widths[i], current_y)
        self.ln()

        for row in data:
            num_lines = [len(wrap(str(row[i]), width=int(col_widths[i] / 2.5))) for i in range(len(row))]
            row_height = max(num_lines) * line_height

            if self.get_y() + row_height > self.page_break_trigger:
                self.add_page()

            y_start = self.get_y()
            x_start = self.get_x()

            for i, item in enumerate(row):
                max_chars = int(col_widths[i] * 0.55)
                lines = wrap(str(item), width=max_chars)
                cell_x = self.get_x()
                cell_y = self.get_y()
                
                self.rect(cell_x, cell_y, col_widths[i], row_height)
                
                text_start_y = cell_y + (row_height - len(lines) * line_height) / 2
                
                for j, line in enumerate(lines):
                    self.set_xy(cell_x, text_start_y + j * line_height)
                    self.cell(col_widths[i], line_height, line, ln=0, align='C')
                
                self.set_xy(x_start + sum(col_widths[:i+1]), y_start)
            self.ln(row_height)

# 🔁 Fonction principale AMÉLIORÉE
def run_search(method, api_key, first_name=None, last_name=None, scopus_id=None, year=None, export_type="pdf", csv_path="scimagojr 2024.csv", debug_mode=False):
    """
    Fonction principale améliorée avec matching intelligent des journaux

    Args:
        api_key (str): Clé API Scopus pour l'authentification
        debug_mode (bool): Active le mode débogage pour les journaux non trouvés
    """

    # Create output directory if it doesn't exist
    output_dir = "output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print("📋 Chargement des données SCImago...")
    scimago_data = load_scimago_data(csv_path)
    print(f"✅ {len(scimago_data)} entrées SCImago chargées")
    
    headers = {
        "X-ELS-APIKey": api_key,
        "Accept": "application/json"
    }

    pdf_filename = None
    excel_filename = None

    if method == "name":
        full_name = f"{last_name} {first_name}".strip()
        query = f'AUTH("{full_name}")'
        if year:
            query += f' AND PUBYEAR IS {year}'
    elif method == "id":
        full_name = scopus_id
        query = f"AU-ID({scopus_id})"
        if year:
            query += f" AND PUBYEAR IS {year}"
        author_info_url = f"https://api.elsevier.com/content/author/author_id/{scopus_id}"
        author_response = requests.get(author_info_url, headers=headers, timeout=10)
        if author_response.status_code == 200:
            author_data = author_response.json()
            author_name = author_data.get("author-retrieval-response", [{}])[0].get("preferred-name", {})
            retrieved_name = f"{author_name.get('surname', '')} {author_name.get('given-name', '')}"
            if retrieved_name.strip():
                full_name = retrieved_name
    else:
        raise ValueError("Invalid search method")

    print(f"🔍 Recherche des publications pour: {full_name}")
    entries = get_all_publications(query, headers)
    print(f"📚 {len(entries)} publications brutes trouvées")

    # 🚨 PREMIÈRE VÉRIFICATION : Si aucune publication trouvée
    if not entries:
        print(f"❌ Aucune publication trouvée pour {full_name}")
        year_text = f" en {year}" if year else ""
        return {
            "data": [],
            "file": {
                "pdf": None,
                "excel": None
            },
            "stats": {
                "total_publications": 0,
                "journals_found": 0,
                "journals_not_found": 0,
                "success_rate": 0
            },
            "message": f"Aucune publication trouvée pour '{full_name}'{year_text}. Vérifiez l'orthographe du nom ou essayez avec l'ID Scopus de l'auteur."
        }

    # 🔍 FILTRAGE : Supprimer les publications sans titre valide
    valid_entries = []
    for pub in entries:
        title = pub.get("dc:title", "").strip()
        # Filtrer les publications sans titre ou avec des titres non valides
        if title and title not in ["No title", "", "N/A", "null", "undefined"]:
            valid_entries.append(pub)

    print(f"📚 {len(valid_entries)} publications valides après filtrage")

    # 🚨 DEUXIÈME VÉRIFICATION : Si aucune publication valide après filtrage
    if not valid_entries:
        print(f"❌ Aucune publication valide trouvée pour {full_name}")
        year_text = f" en {year}" if year else ""
        return {
            "data": [],
            "file": {
                "pdf": None,
                "excel": None
            },
            "stats": {
                "total_publications": 0,
                "journals_found": 0,
                "journals_not_found": 0,
                "success_rate": 0
            },
            "message": f"Aucune publication trouvée pour '{full_name}'{year_text}. Vérifiez l'orthographe du nom ou essayez avec l'ID Scopus de l'auteur."
        }

    # Utiliser les publications valides pour la suite
    entries = valid_entries

    results = []
    journals_not_found = []
    journals_found = []
    
    for i, pub in enumerate(entries, start=1):
        print(f"⏳ Traitement publication {i}/{len(entries)}")

        # Récupération du titre (déjà filtré, donc doit être valide)
        title = pub.get("dc:title", "").strip()
        if not title:  # Sécurité supplémentaire
            title = "Sans titre"

        journal = pub.get("prism:publicationName", "No journal")
        date = pub.get("prism:coverDate", "No date")

        # Extraction du DOI
        doi = pub.get("prism:doi", "-")
        if not doi or doi == "":
            doi = "-"

        citations_count = get_citations_count(pub, headers)
        coauthors = get_coauthors(pub, headers)
        
        document_type = get_document_type(pub)

        # 🔥 NOUVEAU MATCHING INTELLIGENT
        journal_info = improved_journal_matching(journal, scimago_data)
        
        if journal_info:
            journals_found.append(journal)
        else:
            journals_not_found.append(journal)
            if debug_mode:
                debug_journal_matching(journal, scimago_data, csv_path)

        sjr = journal_info.get("sjr", "-")
        h_index = journal_info.get("h_index", "-")
        quartile = journal_info.get("quartile", "-")
        issn = journal_info.get("issn", "-")

        results.append([i, title, journal, date, document_type, citations_count, coauthors, sjr, h_index, quartile, issn, doi])

    # 📊 Rapport final
    print(f"\n📊 RAPPORT FINAL:")
    print(f"✅ Journaux trouvés: {len(journals_found)}/{len(entries)}")
    print(f"❌ Journaux non trouvés: {len(journals_not_found)}/{len(entries)}")
    
    if journals_not_found:
        print(f"\n❌ Journaux non trouvés ({len(journals_not_found)}):")
        for journal in set(journals_not_found):
            print(f"  - {journal}")

    # Create unique filename
    safe_name = re.sub(r'[^\w\-_.]', '_', full_name)
    year_suffix = f"_{year}" if year else "_all_years"
    method_suffix = f"_{method}"
    
    base_filename = f"publications_{safe_name}{year_suffix}{method_suffix}"

    if export_type in ["excel", "both"]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Publications"
        headers_excel = ["N°", "Titre", "Journal", "Date", "Type", "Citations", "Co-auteurs", "SJR", "H-index", "Quartile", "ISSN", "DOI"]
        ws.append(headers_excel)
        for row in results:
            ws.append(row)
        for cell in ws["1:1"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        excel_filename = f"output/{base_filename}.xlsx"
        wb.save(excel_filename)
        print(f"📊 Excel généré: {excel_filename}")

    if export_type in ["pdf", "both"]:
        pdf = PDF()
        pdf.add_page()
        pdf.add_title(full_name, year)
        pdf.render_table(results, ["N°", "Titre", "Journal", "Date", "Type", "Citations", "Co-auteurs", "SJR", "H-index", "Quartile", "ISSN", "DOI"],
                        [8, 50, 40, 18, 25, 15, 30, 12, 12, 12, 18, 25], 4)
        pdf_filename = f"output/{base_filename}.pdf"
        pdf.output(pdf_filename)
        print(f"📄 PDF généré: {pdf_filename}")

    publications_readable = []
    for row in results:
        publications_readable.append({
            "number": row[0],
            "title": row[1],
            "journal": row[2],
            "year": row[3],
            "document_type": row[4],
            "citations": row[5],
            "coauthors": row[6],
            "sjr": row[7],
            "h_index": row[8],
            "quartile": row[9],
            "issn": row[10],
            "doi": row[11],
    })

    print(f"✅ Traitement terminé!")
    
    return {
        "data": publications_readable,
        "file": {
            "pdf": pdf_filename if export_type in ["pdf", "both"] else None,
            "excel": excel_filename if export_type in ["excel", "both"] else None
        },
        "stats": {
            "total_publications": len(entries),
            "journals_found": len(journals_found),
            "journals_not_found": len(journals_not_found),
            "success_rate": round(len(journals_found) / len(entries) * 100, 1) if entries else 0
        }
    }

# Fonction utilitaire pour calculer les statistiques de citations
def calculate_citation_stats(publications_data):
    """Calcule des statistiques sur les citations"""
    citations = []
    for pub in publications_data:
        try:
            citation_count = int(pub['citations']) if pub['citations'] != 'N/A' else 0
            citations.append(citation_count)
        except (ValueError, TypeError):
            citations.append(0)
    
    if not citations:
        return {}
    
    total_citations = sum(citations)
    avg_citations = total_citations / len(citations)
    max_citations = max(citations)
    
    # Calcul de l'indice h simplifié
    citations_sorted = sorted(citations, reverse=True)
    h_index = 0
    for i, citation_count in enumerate(citations_sorted, 1):
        if citation_count >= i:
            h_index = i
        else:
            break
    
    return {
        "total_citations": total_citations,
        "average_citations": round(avg_citations, 2),
        "max_citations": max_citations,
        "h_index": h_index,
        "publications_count": len(citations)
    }
