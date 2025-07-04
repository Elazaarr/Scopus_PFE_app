from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from main import run_search
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from fpdf import FPDF
from textwrap import wrap
import os
import re
import json
from config import Config

app = Flask(__name__)

# Configuration sécurisée
try:
    Config.validate_config()
    app.config.from_object(Config)
    app.secret_key = Config.SECRET_KEY
    UPLOAD_FOLDER = Config.UPLOAD_FOLDER
    print("🔐 Configuration sécurisée chargée avec succès")
except ValueError as e:
    print(f"❌ Erreur de configuration : {e}")
    print("💡 Veuillez vérifier votre fichier .env")
    exit(1)

@app.route('/')
def home():
    return render_template('index.html')

@app.route("/search", methods=['GET','POST'])
def search():
    method = request.args.get("method")
    if request.method == "GET":
        return render_template("search.html", method=method)
    elif request.method == "POST":
        # Récupérer les données du formulaire
        if method == "name":
            name = request.form.get("name")
            # logique avec name
        elif method == "scopus":
            scopus_id = request.form.get("scopus_id")
            # logique avec scopus_id
        # retourne un résultat ou redirige
        return "traitement terminé"

@app.route('/results', methods=['POST'])
def results():
    method = request.form.get('method')
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    scopus_id = request.form.get('scopus_id')
    year = request.form.get('year')
    export_type = request.form.get('export_type')

    try:
        print(f"🔍 Début de recherche: method={method}, first_name={first_name}, last_name={last_name}, year={year}")

        result = run_search(
            method=method,
            api_key=Config.SCOPUS_API_KEY,
            first_name=first_name,
            last_name=last_name,
            scopus_id=scopus_id,
            year=year,
            export_type=export_type
        )

        print(f"📊 Résultat de recherche: {len(result['data'])} publications trouvées")
        print(f"📁 Fichiers générés: PDF={result['file'].get('pdf')}, Excel={result['file'].get('excel')}")

        # Construction du nom à afficher
        if method == 'name':
            name_display = f"{first_name} {last_name}"
        else:
            name_display = f"Scopus ID : {scopus_id}"

        # Récupération des bons fichiers
        pdf_file = result['file'].get('pdf') if 'pdf' in result['file'] else None
        excel_file = result['file'].get('excel') if 'excel' in result['file'] else None

        return render_template('results.html',
                               name=name_display,
                               publications=result['data'],
                               total=len(result['data']),
                               pdf_file=pdf_file,
                               excel_file=excel_file,
                               export_type=export_type)

    except Exception as e:
        print(f"❌ Erreur lors de la recherche: {str(e)}")
        flash(f"Erreur : {str(e)}", 'danger')
        return redirect(url_for('home'))

@app.route('/download/<path:filename>')
def download_file(filename):
    # Remove 'output/' from filename if it exists
    if filename.startswith('output/'):
        filename = filename[7:]  # Remove 'output/' prefix

    file_path = os.path.join(UPLOAD_FOLDER, 'output', filename)

    # Check if file exists
    if not os.path.exists(file_path):
        flash(f"File not found: {filename}", 'danger')
        return redirect(url_for('home'))

    return send_file(file_path, as_attachment=True)

# Classe PDF personnalisée pour les téléchargements filtrés
class FilteredPDF(FPDF):
    def __init__(self):
        super().__init__(orientation='L', unit='mm', format='A4')
        self.set_auto_page_break(auto=True, margin=10)
        try:
            self.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
            self.set_font('DejaVu', size=8)
        except:
            self.set_font('Arial', size=8)

    def add_title(self, author_name, filters_info):
        """Ajoute un titre formaté avec les informations de filtrage"""
        try:
            self.set_font('DejaVu', size=16)
        except:
            self.set_font('Arial', size=16)

        title_line1 = f"Publications filtrées de : {author_name}"

        self.set_xy(0, 20)
        self.cell(0, 10, title_line1, ln=True, align='C')

        # Ajouter les informations de filtrage
        if filters_info:
            try:
                self.set_font('DejaVu', size=10)
            except:
                self.set_font('Arial', size=10)

            self.ln(5)
            filter_text = "Filtres appliqués: "
            active_filters = []

            if filters_info.get('title'):
                active_filters.append(f"Titre: {filters_info['title']}")
            if filters_info.get('journal'):
                active_filters.append(f"Journal: {filters_info['journal']}")
            if filters_info.get('coauthor'):
                active_filters.append(f"Co-auteur: {filters_info['coauthor']}")
            if filters_info.get('year'):
                active_filters.append(f"Année: {filters_info['year']}")
            if filters_info.get('quartile'):
                active_filters.append(f"Quartile: {filters_info['quartile']}")

            if active_filters:
                filter_text += " | ".join(active_filters)
            else:
                filter_text += "Aucun filtre appliqué"

            self.set_x(0)
            self.cell(0, 8, filter_text, ln=True, align='C')

        self.ln(15)

        try:
            self.set_font('DejaVu', size=8)
        except:
            self.set_font('Arial', size=8)

    def render_table(self, data, headers, col_widths, line_height):
        try:
            self.set_font('DejaVu', size=8)
        except:
            self.set_font('Arial', size=8)

        # En-têtes
        for i, header in enumerate(headers):
            self.set_xy(self.get_x(), self.get_y())
            self.cell(col_widths[i], 8, header, border=1, align='C')
            current_x = self.get_x() - col_widths[i]
            current_y = self.get_y()
            self.set_xy(current_x + 0.1, current_y)
            self.cell(col_widths[i], 8, header, border=0, align='C')
            self.set_xy(current_x + col_widths[i], current_y)
        self.ln()

        # Données
        for row in data:
            num_lines = [len(wrap(str(row[i]), width=int(col_widths[i] / 2.5))) for i in range(len(row))]
            row_height = max(num_lines) * line_height

            if self.get_y() + row_height > self.page_break_trigger:
                self.add_page()

            y_start = self.get_y()
            x_start = self.get_x()

            for i, item in enumerate(row):
                max_chars = int(col_widths[i] * 0.55)
                lines = wrap(str(item), width=max_chars)
                cell_x = self.get_x()
                cell_y = self.get_y()

                self.rect(cell_x, cell_y, col_widths[i], row_height)

                text_start_y = cell_y + (row_height - len(lines) * line_height) / 2

                for j, line in enumerate(lines):
                    self.set_xy(cell_x, text_start_y + j * line_height)
                    self.cell(col_widths[i], line_height, line, ln=0, align='C')

                self.set_xy(x_start + sum(col_widths[:i+1]), y_start)
            self.ln(row_height)

@app.route('/download_filtered_pdf', methods=['POST'])
def download_filtered_pdf():
    try:
        # Récupérer les données du formulaire
        filtered_data_json = request.form.get('filtered_data')
        author_name = request.form.get('author_name')
        applied_filters_json = request.form.get('applied_filters')

        if not filtered_data_json:
            flash('Aucune donnée à télécharger', 'danger')
            return redirect(url_for('home'))

        # Parser les données JSON
        filtered_data = json.loads(filtered_data_json)
        applied_filters = json.loads(applied_filters_json) if applied_filters_json else {}

        if not filtered_data:
            flash('Aucune donnée filtrée disponible', 'danger')
            return redirect(url_for('home'))

        # Créer le répertoire de sortie s'il n'existe pas
        output_dir = os.path.join(UPLOAD_FOLDER, 'output')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Préparer les données pour le PDF (sans colonne N°)
        pdf_data = []
        for pub in filtered_data:
            pdf_data.append([
                pub.get('title', 'Sans titre'),
                pub.get('journal', 'Journal non spécifié'),
                pub.get('document_type', 'Non spécifié'),
                pub.get('coauthors', 'Non spécifiés'),
                pub.get('year', '-'),
                pub.get('citations', '0'),
                pub.get('sjr', '-'),
                pub.get('h_index', '-'),
                pub.get('quartile', '-'),
                pub.get('issn', '-'),
                pub.get('doi', '-')
            ])

        # Créer le PDF
        pdf = FilteredPDF()
        pdf.add_page()
        pdf.add_title(author_name, applied_filters)
        pdf.render_table(
            pdf_data,
            ["Titre", "Journal", "Type", "Co-auteurs", "Année", "Citations", "SJR", "H-index", "Quartile", "ISSN", "DOI"],
            [50, 35, 25, 18, 12, 12, 8, 8, 9, 12, 20],
            4
        )

        # Générer un nom de fichier unique
        safe_name = re.sub(r'[^\w\-_.]', '_', author_name)
        filename = f"publications_filtrees_{safe_name}_{len(filtered_data)}_resultats.pdf"
        file_path = os.path.join(output_dir, filename)

        pdf.output(file_path)

        return send_file(file_path, as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'Erreur lors de la génération du PDF: {str(e)}', 'danger')
        return redirect(url_for('home'))

@app.route('/download_filtered_excel', methods=['POST'])
def download_filtered_excel():
    try:
        # Récupérer les données du formulaire
        filtered_data_json = request.form.get('filtered_data')
        author_name = request.form.get('author_name')
        applied_filters_json = request.form.get('applied_filters')

        if not filtered_data_json:
            flash('Aucune donnée à télécharger', 'danger')
            return redirect(url_for('home'))

        # Parser les données JSON
        filtered_data = json.loads(filtered_data_json)
        applied_filters = json.loads(applied_filters_json) if applied_filters_json else {}

        if not filtered_data:
            flash('Aucune donnée filtrée disponible', 'danger')
            return redirect(url_for('home'))

        # Créer le répertoire de sortie s'il n'existe pas
        output_dir = os.path.join(UPLOAD_FOLDER, 'output')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Créer le classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Publications Filtrées"

        # Ajouter les informations de filtrage en en-tête
        ws.append([f"Publications filtrées de : {author_name}"])

        # Ajouter les filtres appliqués
        filter_info = []
        if applied_filters.get('title'):
            filter_info.append(f"Titre: {applied_filters['title']}")
        if applied_filters.get('journal'):
            filter_info.append(f"Journal: {applied_filters['journal']}")
        if applied_filters.get('coauthor'):
            filter_info.append(f"Co-auteur: {applied_filters['coauthor']}")
        if applied_filters.get('year'):
            filter_info.append(f"Année: {applied_filters['year']}")
        if applied_filters.get('quartile'):
            filter_info.append(f"Quartile: {applied_filters['quartile']}")

        if filter_info:
            ws.append([f"Filtres appliqués: {' | '.join(filter_info)}"])
        else:
            ws.append(["Aucun filtre appliqué"])

        ws.append([])  # Ligne vide

        # En-têtes des colonnes (sans N°)
        headers_excel = ["Titre", "Journal", "Type", "Co-auteurs", "Année", "Citations", "SJR", "H-index", "Quartile", "ISSN", "DOI"]
        ws.append(headers_excel)

        # Données filtrées
        for pub in filtered_data:
            row_data = [
                pub.get('title', 'Sans titre'),
                pub.get('journal', 'Journal non spécifié'),
                pub.get('document_type', 'Non spécifié'),
                pub.get('coauthors', 'Non spécifiés'),
                pub.get('year', '-'),
                pub.get('citations', '0'),
                pub.get('sjr', '-'),
                pub.get('h_index', '-'),
                pub.get('quartile', '-'),
                pub.get('issn', '-'),
                pub.get('doi', '-')
            ]
            ws.append(row_data)

        # Formatage des en-têtes
        header_row = ws[4]  # La ligne des en-têtes (après les infos et ligne vide)
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Formatage des informations en haut
        ws[1][0].font = Font(bold=True, size=14)
        ws[2][0].font = Font(italic=True)

        # Générer un nom de fichier unique
        safe_name = re.sub(r'[^\w\-_.]', '_', author_name)
        filename = f"publications_filtrees_{safe_name}_{len(filtered_data)}_resultats.xlsx"
        file_path = os.path.join(output_dir, filename)

        wb.save(file_path)

        return send_file(file_path, as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'Erreur lors de la génération du fichier Excel: {str(e)}', 'danger')
        return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True)